"""Excel parser for 基金E账户App exported Excel files.

Expected format:
- Row 1: Title
- Row 2-3: User info
- Row 4: Empty
- Row 5: Column headers
- Row 6+: Data rows
"""

from __future__ import annotations

import hashlib
from zipfile import BadZipFile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# Expected column mapping (1-indexed)
COLUMN_MAP = {
    "序号": 1,
    "基金代码": 2,
    "基金名称": 3,
    "份额类别": 4,
    "基金管理人": 5,
    "基金账户": 6,
    "销售机构": 7,
    "交易账户": 8,
    "持有份额": 9,
    "份额日期": 10,
    "基金净值": 11,
    "净值日期": 12,
    "资产情况": 13,
    "结算币种": 14,
    "分红方式": 15,
}


class ExcelParseError(Exception):
    """Raised when Excel parsing fails."""
    pass


class BatchValidationError(ExcelParseError):
    """Raised when a whole upload must be rejected before database writes."""
    pass


class ParsedHolding:
    """Parsed holding record from Excel."""

    def __init__(
        self,
        fund_code: str,
        fund_name: str,
        share_type: str,
        management_company: str,
        fund_account: str,
        platform: str,
        trade_account: str,
        shares: Decimal,
        share_date: date,
        nav: Optional[Decimal],
        nav_date: Optional[date],
        market_value: Optional[Decimal],
        currency: str,
        dividend_mode: Optional[str],
    ):
        self.fund_code = fund_code
        self.fund_name = fund_name
        self.share_type = share_type
        self.management_company = management_company
        self.fund_account = fund_account
        self.platform = platform
        self.trade_account = trade_account
        self.shares = shares
        self.share_date = share_date
        self.nav = nav
        self.nav_date = nav_date
        self.market_value = market_value
        self.currency = currency
        self.dividend_mode = dividend_mode

    @property
    def unique_key(self) -> tuple:
        return (self.fund_code, self.platform, self.fund_account, self.trade_account)


def _parse_date(value: str | None) -> Optional[date]:
    """Parse date string in various formats."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    value = str(value).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: str | None) -> Optional[Decimal]:
    """Parse decimal value from string."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return Decimal(str(value).strip().replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def compute_file_hash(file_path: str | Path) -> str:
    """Compute SHA256 hash of a file."""
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def validate_headers(ws) -> bool:
    """Validate that row 5 contains expected column headers."""
    row5 = [cell.value for cell in ws[5]]
    expected = list(COLUMN_MAP)
    for i, exp in enumerate(expected):
        actual = str(row5[i]).strip() if row5[i] else ""
        if actual != exp:
            raise ExcelParseError(
                f"列头验证失败: 第{i+1}列期望'{exp}', 实际'{actual}'"
            )
    return True


def parse_excel(file_path: str | Path) -> tuple[list[ParsedHolding], list[dict], Optional[date]]:
    """Parse fund holdings Excel file.

    Returns:
        (holdings, errors, data_date)
        - holdings: list of successfully parsed holding records
        - errors: list of error dicts with row number and message
        - data_date: the common date from the data, if consistent
    """
    path = Path(file_path)
    if path.suffix.lower() == ".xls":
        raise ExcelParseError("不支持 .xls 格式，请转换为 .xlsx")
    if path.suffix.lower() != ".xlsx":
        raise ExcelParseError("只支持 .xlsx 格式")

    try:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ExcelParseError("xlsx 文件损坏或无法读取") from exc
    try:
        holdings: list[ParsedHolding] = []
        errors: list[dict] = []
        data_dates: set[date] = set()
        valid_sheet_count = 0

        for worksheet in workbook.worksheets:
            if _sheet_is_blank(worksheet):
                continue
            try:
                validate_headers(worksheet)
            except ExcelParseError as exc:
                if _sheet_has_business_rows(worksheet):
                    errors.append({"sheet": worksheet.title, "row": 5, "message": str(exc)})
                continue
            valid_sheet_count += 1
            sheet_holdings, sheet_errors, sheet_dates = _parse_sheet(worksheet)
            holdings.extend(sheet_holdings)
            errors.extend({**error, "sheet": worksheet.title} for error in sheet_errors)
            data_dates.update(sheet_dates)

        if valid_sheet_count == 0:
            raise ExcelParseError("未找到可识别的持仓工作表")
        if len(data_dates) > 1:
            raise BatchValidationError("同一上传包含多个业务日期，拒绝合并")
        if not holdings:
            raise ExcelParseError("未找到有效持仓记录")

        return _deduplicate_holdings(holdings), errors, next(iter(data_dates), None)
    finally:
        workbook.close()


def _sheet_is_blank(worksheet) -> bool:
    """判断工作表是否完全为空。"""
    return all(cell.value is None for row in worksheet.iter_rows() for cell in row)


def _sheet_has_business_rows(worksheet) -> bool:
    """判断工作表是否包含疑似业务数据行。"""
    for row in worksheet.iter_rows(min_row=6, values_only=True):
        if any(value not in (None, "") for value in row[:15]):
            return True
    return False


def _parse_sheet(worksheet) -> tuple[list[ParsedHolding], list[dict], set[date]]:
    """解析一个已通过表头校验的工作表。"""
    holdings: list[ParsedHolding] = []
    errors: list[dict] = []
    data_dates: set[date] = set()
    for row_idx in range(6, worksheet.max_row + 1):
        row = [worksheet.cell(row=row_idx, column=c).value for c in range(1, 16)]
        seq = row[0]
        if seq is None:
            continue
        try:
            float(str(seq))
        except ValueError:
            continue
        if isinstance(row[1], (int, float)):
            fund_code = f"{int(row[1]):06d}"
        else:
            fund_code = str(row[1]).strip() if row[1] else None
        if not fund_code:
            errors.append({"row": row_idx, "message": "基金代码为空"})
            continue
        shares = _parse_decimal(row[8])
        if shares is None:
            errors.append({"row": row_idx, "message": f"份额解析失败: {row[8]}"})
            continue
        share_date = _parse_date(row[9])
        if share_date is None:
            errors.append({"row": row_idx, "message": f"份额日期解析失败: {row[9]}"})
            continue
        data_dates.add(share_date)
        holdings.append(ParsedHolding(
            fund_code=fund_code,
            fund_name=str(row[2]).strip() if row[2] else "",
            share_type=str(row[3]).strip() if row[3] else "前收费",
            management_company=str(row[4]).strip() if row[4] else "",
            fund_account=str(row[5]).strip() if row[5] else "",
            platform=str(row[6]).strip() if row[6] else "",
            trade_account=str(row[7]).strip() if row[7] else "",
            shares=shares,
            share_date=share_date,
            nav=_parse_decimal(row[10]),
            nav_date=_parse_date(row[11]),
            market_value=_parse_decimal(row[12]),
            currency=str(row[13]).strip() if row[13] else "人民币",
            dividend_mode=str(row[14]).strip() if row[14] else None,
        ))
    return holdings, errors, data_dates


def _deduplicate_holdings(holdings: list[ParsedHolding]) -> list[ParsedHolding]:
    """稳定去重完全相同记录，并拒绝同键内容冲突。"""
    unique: dict[tuple, ParsedHolding] = {}
    for holding in holdings:
        previous = unique.get(holding.unique_key)
        if previous is None:
            unique[holding.unique_key] = holding
            continue
        current_values = vars(holding).copy()
        previous_values = vars(previous).copy()
        if current_values != previous_values:
            raise BatchValidationError(f"唯一键重复且内容冲突: {holding.fund_code}/{holding.platform}")
    return list(unique.values())

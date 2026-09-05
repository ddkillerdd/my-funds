"""Stage C 文件导入安全与批次语义回归测试。"""

from datetime import date
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
import asyncio
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from backend.schemas.import_result import ImportResult
from backend.services.excel_parser import BatchValidationError, ExcelParseError, parse_excel
from backend.services.import_service import ImportService


HEADERS = [
    "序号", "基金代码", "基金名称", "份额类别", "基金管理人", "基金账户",
    "销售机构", "交易账户", "持有份额", "份额日期", "基金净值", "净值日期",
    "资产情况", "结算币种", "分红方式",
]


def _write_book(path: Path, sheets: list[tuple[str, list[list[object]]]]) -> None:
    """创建虚构的多 Sheet 持仓工作簿。"""
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets:
        sheet = workbook.create_sheet(title)
        if title == "说明":
            sheet.append(["这是不含业务表头的说明页"])
            continue
        sheet.append(["说明页"])
        for row in range(2, 5):
            sheet.cell(row=row, column=1, value=None)
        for column, value in enumerate(HEADERS, 1):
            sheet.cell(row=5, column=column, value=value)
        for values in rows:
            sheet.append(values)
    workbook.save(path)


def _holding_row(code: str, platform: str, share_date: date, shares: str = "1") -> list[object]:
    """构造一行虚构持仓。"""
    return [1, code, "虚构基金", "前收费", "虚构管理人", "账户", platform, "交易", shares, share_date, "2", share_date, "2", "人民币", None]


def test_xls_is_rejected_before_openpyxl(tmp_path):
    """验证独立 .xls 被明确拒绝。"""
    with pytest.raises(ExcelParseError, match="不支持 .xls"):
        parse_excel(tmp_path / "holding.xls")


def test_multisheet_parser_merges_and_closes_workbook(tmp_path, monkeypatch):
    """验证多个有效 Sheet 合并、空白说明页忽略且工作簿最终关闭。"""
    path = tmp_path / "holding.xlsx"
    _write_book(path, [("持仓", [_holding_row("000001", "平台A", date(2026, 9, 4))]), ("说明", [])])
    loaded = []
    original = load_workbook

    def tracked(*args, **kwargs):
        workbook = original(*args, **kwargs)
        loaded.append(workbook)
        original_close = workbook.close

        def close_and_record():
            closed.append(True)
            return original_close()

        workbook.close = close_and_record
        return workbook

    closed = []
    monkeypatch.setattr("backend.services.excel_parser.openpyxl.load_workbook", tracked)
    holdings, errors, data_date = parse_excel(path)
    assert len(holdings) == 1
    assert errors == []
    assert data_date == date(2026, 9, 4)
    assert loaded and closed == [True]


def test_mixed_dates_are_rejected_before_merge(tmp_path):
    """验证跨 Sheet 混合业务日期整批拒绝。"""
    path = tmp_path / "mixed.xlsx"
    _write_book(path, [("一", [_holding_row("000001", "平台A", date(2026, 9, 4))]), ("二", [_holding_row("000002", "平台B", date(2026, 9, 5))])])
    with pytest.raises(BatchValidationError, match="多个业务日期"):
        parse_excel(path)


def test_duplicate_rows_dedupe_and_conflict_reject(tmp_path):
    """验证完全相同唯一键稳定去重，内容冲突拒绝。"""
    same = _holding_row("000001", "平台A", date(2026, 9, 4))
    path = tmp_path / "duplicate.xlsx"
    _write_book(path, [("一", [same]), ("二", [same.copy()])])
    holdings, _, _ = parse_excel(path)
    assert len(holdings) == 1
    conflict = same.copy()
    conflict[8] = "2"
    _write_book(path, [("一", [same]), ("二", [conflict])])
    with pytest.raises(BatchValidationError, match="重复且内容冲突"):
        parse_excel(path)


def test_zip_xls_member_is_rejected_without_extractall(tmp_path):
    """验证 ZIP 内 .xls 被拒绝。"""
    zip_path = tmp_path / "holdings.zip"
    with ZipFile(zip_path, "w", ZIP_DEFLATED) as archive:
        archive.writestr("nested/old.xls", b"not-an-xls")
    with ZipFile(zip_path) as archive:
        member = archive.infolist()[0]
        with pytest.raises(ExcelParseError, match="不支持 .xls"):
            ImportService(None)._validate_zip_members([member])


def test_zip_path_validation_rejects_slashes_drive_unc_and_nul():
    """验证 ZIP 路径穿越、盘符、UNC、反斜杠和 NUL 均拒绝。"""
    service = ImportService(None)
    for name in ("../x.xlsx", r"..\x.xlsx", "/x.xlsx", r"\x.xlsx", "C:x.xlsx", r"\\server\x.xlsx", "x\x00.xlsx"):
        member = SimpleNamespace(filename=name, file_size=1, external_attr=0, flag_bits=0, is_dir=lambda: False)
        with pytest.raises(ValueError):
            service._validate_zip_members([member])


def test_zip_limits_cover_member_count_and_sizes():
    """验证成员数量、单成员大小和总解压大小门禁。"""
    service = ImportService(None)
    make = lambda name, size: SimpleNamespace(filename=name, file_size=size, external_attr=0, flag_bits=0, is_dir=lambda: False)
    with pytest.raises(ValueError, match="成员数"):
        service._validate_zip_members([make(f"member-{index}.xlsx", 1) for index in range(101)])
    with pytest.raises(ValueError, match="单个文件"):
        service._validate_zip_members([make("single.xlsx", 20 * 1024 * 1024 + 1)])
    with pytest.raises(ValueError, match="总量"):
        service._validate_zip_members([
            make(f"member-{index}.xlsx", 20 * 1024 * 1024) for index in range(6)
        ])


def test_zip_rejects_duplicate_casefolded_members_and_special_members():
    """验证重复/大小写碰撞、符号链接、加密和非普通成员均拒绝。"""
    service = ImportService(None)
    make = lambda name, attr=0, flags=0: SimpleNamespace(filename=name, file_size=1, external_attr=attr, flag_bits=flags, is_dir=lambda: False)
    with pytest.raises(ValueError, match="重复"):
        service._validate_zip_members([make("A.xlsx"), make("a.xlsx")])
    with pytest.raises(ValueError, match="符号链接"):
        service._validate_zip_members([make("link.xlsx", 0o120000 << 16)])
    with pytest.raises(ValueError, match="加密"):
        service._validate_zip_members([make("secret.xlsx", flags=1)])
    with pytest.raises(ValueError, match="普通文件"):
        service._validate_zip_members([make("device.xlsx", 0o060000 << 16)])


def test_decimal_parser_preserves_zero_values():
    """验证零、零浮点和 Decimal 零不是空值。"""
    from backend.services.excel_parser import _parse_decimal

    assert _parse_decimal(0) == Decimal("0")
    assert _parse_decimal(0.0) == Decimal("0.0")
    assert _parse_decimal(Decimal("0")) == Decimal("0")


class _Upload:
    """提供分块读取的虚构上传对象。"""

    def __init__(self, filename, content):
        self.filename = filename
        self.content = BytesIO(content)

    async def read(self, size=-1):
        return self.content.read(size)


def test_upload_empty_and_oversize_are_rejected(tmp_path):
    """验证空上传和超过 20 MiB 上传均拒绝。"""
    service = ImportService(None)
    with pytest.raises(ValueError, match="不能为空"):
        asyncio.run(service._save_upload(_Upload("safe.xlsx", b""), tmp_path / "empty"))
    with pytest.raises(ValueError, match="超过 20MB"):
        asyncio.run(service._save_upload(_Upload("safe.xlsx", b"x" * (20 * 1024 * 1024 + 1)), tmp_path / "large"))


def test_zip_two_files_are_one_batch_and_one_commit(monkeypatch, tmp_path):
    """验证 ZIP 两个文件共同进入一次 merge/commit，且两份持仓同时保留。"""
    zip_bytes = BytesIO()
    with ZipFile(zip_bytes, "w", ZIP_DEFLATED) as archive:
        archive.writestr("z/second.xlsx", b"second")
        archive.writestr("first.xlsx", b"first")
    parsed = [
        SimpleNamespace(unique_key=("000001", "平台A"), fund_code="000001"),
        SimpleNamespace(unique_key=("000002", "平台B"), fund_code="000002"),
    ]
    parse_calls = []
    monkeypatch.setattr("backend.services.import_service.parse_excel", lambda path: (parse_calls.append(path.name) or ([parsed[len(parse_calls) - 1]], [], date(2026, 9, 4))))
    class Db:
        """记录最终导入记录与事务次数。"""
        commits = 0
        refreshes = 0
        def execute(self, _statement):
            return SimpleNamespace(scalar_one_or_none=lambda: None)
        def add(self, record):
            if hasattr(record, "file_hash"):
                record.id = 77
        def flush(self):
            pass
        def commit(self):
            self.commits += 1
        def refresh(self, _record):
            self.refreshes += 1
    db = Db()
    service = ImportService(db)
    merge_calls = []
    service._merge_holdings = lambda holdings, import_id, data_date, allow_clear=True: (merge_calls.append((holdings, import_id, allow_clear)) or (2, 0, 0, []))
    result = asyncio.run(service.import_zip(_Upload("C:\\client\\holdings.zip", zip_bytes.getvalue())))
    assert result.import_id == 77
    assert len(parse_calls) == 2
    assert len(merge_calls) == 1
    assert {item.fund_code for item in merge_calls[0][0]} == {"000001", "000002"}
    assert merge_calls[0][1] == result.import_id and merge_calls[0][2] is True
    assert db.commits == 1
    assert db.refreshes == 0


def test_batch_flush_merge_and_commit_failures_rollback():
    """验证 flush、merge、commit 任一失败都会 rollback。"""
    class Db:
        def __init__(self, fail_at):
            self.fail_at = fail_at
            self.rollbacks = 0
        def add(self, record):
            record.id = 1
        def flush(self):
            if self.fail_at == "flush":
                raise RuntimeError("flush failed")
        def commit(self):
            if self.fail_at == "commit":
                raise RuntimeError("commit failed")
        def refresh(self, _record):
            pass
        def rollback(self):
            self.rollbacks += 1
    for failure in ("flush", "merge", "commit"):
        db = Db(failure)
        service = ImportService(db)
        if failure == "merge":
            service._merge_holdings = lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("merge failed"))
        elif failure == "commit":
            service._merge_holdings = lambda *args, **kwargs: (0, 0, 0, [])
        with pytest.raises(RuntimeError):
            service._import_parsed_batch("safe.xlsx", "hash", [SimpleNamespace()], [], date(2026, 9, 4))
        assert db.rollbacks == 1


def test_merge_clear_protects_non_file_and_missing_import_sources():
    """验证清除只针对 file 且具有 last_import_id 的持仓。"""
    active = [
        SimpleNamespace(id=1, fund_code="F1", platform="P", fund_account="A", trade_account="T", source_type="manual", last_import_id=1, status=1, shares=Decimal("1"), market_value=Decimal("1"), fund_name="F1", nav_on_import=Decimal("1")),
        SimpleNamespace(id=2, fund_code="F2", platform="P", fund_account="A", trade_account="T", source_type="quick", last_import_id=2, status=1, shares=Decimal("1"), market_value=Decimal("1"), fund_name="F2", nav_on_import=Decimal("1")),
        SimpleNamespace(id=3, fund_code="F3", platform="P", fund_account="A", trade_account="T", source_type="legacy", last_import_id=3, status=1, shares=Decimal("1"), market_value=Decimal("1"), fund_name="F3", nav_on_import=Decimal("1")),
        SimpleNamespace(id=4, fund_code="F4", platform="P", fund_account="A", trade_account="T", source_type="file", last_import_id=None, status=1, shares=Decimal("1"), market_value=Decimal("1"), fund_name="F4", nav_on_import=Decimal("1")),
        SimpleNamespace(id=5, fund_code="F5", platform="P", fund_account="A", trade_account="T", source_type="file", last_import_id=5, status=1, shares=Decimal("1"), market_value=Decimal("1"), fund_name="F5", nav_on_import=Decimal("1")),
    ]
    class Db:
        def execute(self, _statement):
            return SimpleNamespace(scalars=lambda: SimpleNamespace(all=lambda: active))
        def add(self, _value):
            pass
        def flush(self):
            pass
    service = ImportService(Db())
    _, _, removed, _ = service._merge_holdings([], 8, date(2026, 9, 4))
    assert removed == 1
    assert all(item.status == 1 for item in active[:4])
    assert active[4].status == 0


def test_zip_actual_total_limit_and_cleanup(monkeypatch, tmp_path):
    """验证实际解压总量超过 100 MiB 时拒绝，并清理上传文件和临时目录。"""
    archive_bytes = BytesIO()
    with ZipFile(archive_bytes, "w", ZIP_DEFLATED) as archive:
        for index in range(6):
            archive.writestr(f"file{index}.xlsx", b"x")

    class Db:
        def execute(self, _statement):
            return SimpleNamespace(scalar_one_or_none=lambda: None)

    monkeypatch.setattr("backend.services.import_service.UPLOAD_DIR", tmp_path)
    monkeypatch.setattr("backend.services.import_service.ImportService._copy_zip_member", lambda *args: 20 * 1024 * 1024)
    result = asyncio.run(ImportService(Db()).import_zip(_Upload("holdings.zip", archive_bytes.getvalue())))
    assert result.status == "error"
    assert "实际解压总量" in result.error_message
    assert list(tmp_path.iterdir()) == []


def test_real_explanation_sheet_is_skipped_and_bad_business_sheet_is_partial(tmp_path):
    """验证真实说明页安全跳过，错误业务 Sheet 保留诊断。"""
    path = tmp_path / "sheets.xlsx"
    _write_book(path, [("有效", [_holding_row("000001", "平台A", date(2026, 9, 4))]), ("说明", [])])
    workbook = load_workbook(path)
    bad = workbook.create_sheet("错误业务")
    bad.cell(row=6, column=1, value=1)
    bad.cell(row=6, column=2, value="000002")
    workbook.save(path)
    holdings, errors, _ = parse_excel(path)
    assert len(holdings) == 1
    assert any(error.get("sheet") == "错误业务" for error in errors)


def test_corrupt_xlsx_returns_controlled_error_and_cleans_upload(monkeypatch, tmp_path):
    """验证损坏 xlsx 转受控错误，并清理上传临时文件。"""
    monkeypatch.setattr("backend.services.import_service.UPLOAD_DIR", tmp_path)
    db = SimpleNamespace(execute=lambda _statement: SimpleNamespace(scalar_one_or_none=lambda: None))
    result = asyncio.run(ImportService(db).import_excel(_Upload("corrupt.xlsx", b"not-a-zip")))
    assert result.status == "error"
    assert "损坏" in result.error_message
    assert list(tmp_path.iterdir()) == []


def test_zip_batch_validation_error_never_calls_merge(monkeypatch, tmp_path):
    """验证 ZIP 子文件冲突等批次级校验错误不会降级为 partial 或调用 merge。"""
    archive_bytes = BytesIO()
    with ZipFile(archive_bytes, "w", ZIP_DEFLATED) as archive:
        archive.writestr("first.xlsx", b"first")
        archive.writestr("second.xlsx", b"second")
    monkeypatch.setattr("backend.services.import_service.UPLOAD_DIR", tmp_path)
    monkeypatch.setattr("backend.services.import_service.parse_excel", lambda _path: (_ for _ in ()).throw(BatchValidationError("重复且内容冲突")))
    db = SimpleNamespace(execute=lambda _statement: SimpleNamespace(scalar_one_or_none=lambda: None))
    service = ImportService(db)
    calls = []
    service._import_parsed_batch = lambda *args: calls.append(args)
    result = asyncio.run(service.import_zip(_Upload("holdings.zip", archive_bytes.getvalue())))
    assert result.status == "error"
    assert calls == []
    assert list(tmp_path.iterdir()) == []


def test_parse_and_database_errors_clean_upload_artifacts(monkeypatch, tmp_path):
    """验证解析异常和数据库异常路径都清理上传文件及临时对象。"""
    monkeypatch.setattr("backend.services.import_service.UPLOAD_DIR", tmp_path)
    monkeypatch.setattr("backend.services.import_service.parse_excel", lambda _path: (_ for _ in ()).throw(ExcelParseError("解析失败")))
    db = SimpleNamespace(execute=lambda _statement: SimpleNamespace(scalar_one_or_none=lambda: None))
    result = asyncio.run(ImportService(db).import_excel(_Upload("holdings.xlsx", b"content")))
    assert result.status == "error"
    assert list(tmp_path.iterdir()) == []

    class BrokenDb:
        def execute(self, _statement):
            raise RuntimeError("db failed")
    with pytest.raises(RuntimeError):
        asyncio.run(ImportService(BrokenDb()).import_excel(_Upload("holdings.xlsx", b"content")))
    assert list(tmp_path.iterdir()) == []


def test_partial_merge_disables_clear(monkeypatch):
    """验证 partial 批次向合并入口传递 allow_clear=False。"""
    calls = []
    service = ImportService(SimpleNamespace())
    service.db.add = lambda record: setattr(record, "id", 9)
    service.db.flush = lambda: None
    service.db.commit = lambda: None
    service.db.refresh = lambda record: None
    service._merge_holdings = lambda holdings, import_id, data_date, allow_clear=True: (calls.append(allow_clear) or (1, 0, 0, []))
    result = service._import_parsed_batch("safe.xlsx", "hash", [SimpleNamespace()], [{"row": 6, "message": "坏行"}], date(2026, 9, 4))
    assert calls == [False]
    assert result.status == "partial"


def test_import_result_changes_lists_are_not_shared():
    """验证 ImportResult 实例不共享 changes 列表。"""
    first = ImportResult(import_id=1, file_name="a.xlsx")
    second = ImportResult(import_id=2, file_name="b.xlsx")
    first.changes.append("虚构")
    assert second.changes == []


def test_error_text_does_not_contain_server_path():
    """验证文件错误文本只使用展示名，不泄露服务器绝对路径。"""
    result = ImportResult(import_id=0, file_name="safe.xlsx", status="error", error_message="safe.xlsx: 解析失败")
    assert "E:\\" not in result.error_message
    assert "/srv/" not in result.error_message
